import openpyxl
from collections import defaultdict
from datetime import datetime
import unicodedata
import re
import pandas as pd
import os
import streamlit as st
from tempfile import NamedTemporaryFile

# === OUTILS ===
def nettoyer_mois(mois_texte):
    mois_texte = str(mois_texte).lower()
    mois_texte = unicodedata.normalize('NFD', mois_texte).encode('ascii', 'ignore').decode("utf-8")
    match = re.search(r"(janvier|fevrier|mars|avril|mai|juin|juillet|aout|septembre|octobre|novembre|decembre)[^\d]*(\d{4})", mois_texte)
    if not match:
        return None, None
    mois_str, annee = match.groups()
    mois_map = {"janvier":1, "fevrier":2, "mars":3, "avril":4, "mai":5, "juin":6,
                "juillet":7, "aout":8, "septembre":9, "octobre":10, "novembre":11, "decembre":12}
    return mois_map.get(mois_str), int(annee)

def charger_cours(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    cours = []
    feuilles = wb.sheetnames

    for sheet_name in feuilles:
        ws = wb[sheet_name]
        mois_texte = ws["B1"].value or ""
        mois, annee = nettoyer_mois(mois_texte)
        if not mois:
            continue

        classes = {
            "BAC PRO 22": 4,
            "BAC PRO 23": 6,
            "BAC PRO 24": 8,
            "BAC PRO 25": 10,
            "CORA 1 et 2": 12,
            "EC 2": 14
        }

        lignes_saut = {16, 27, 38}
        semaines = [range(6, 16, 2), range(17, 27, 2), range(28, 38, 2), range(39, 49, 2)]

        for semaine in semaines:
            for ligne in semaine:
                if ligne in lignes_saut:
                    continue
                jour_val = ws.cell(row=ligne, column=2).value
                if not jour_val:
                    continue
                match = re.search(r"(\d{1,2})", str(jour_val))
                if not match:
                    continue
                jour = int(match.group(1))
                try:
                    date = datetime(annee, mois, jour).strftime("%d/%m/%Y")
                except:
                    continue
                for classe, col in classes.items():
                    matin = str(ws.cell(row=ligne, column=col).value).strip().upper()
                    aprem = str(ws.cell(row=ligne+1, column=col).value).strip().upper()
                    if matin == "X":
                        cours.append((date, "AM", classe, sheet_name, ligne, col))
                    if aprem == "X":
                        cours.append((date, "PM", classe, sheet_name, ligne+1, col))

    wb.close()
    return cours

def charger_dispos(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    dispo = defaultdict(lambda: defaultdict(set))
    for nom in wb.sheetnames:
        ws = wb[nom]
        for col in range(1, ws.max_column, 2):
            mois_cell = ws.cell(row=5, column=col).value
            if not isinstance(mois_cell, datetime):
                continue
            matin_col, am_col = col, col + 1
            for row in range(7, ws.max_row + 1):
                try:
                    jour = int(str(ws.cell(row=row, column=2).value))
                    full_date = datetime(mois_cell.year, mois_cell.month, jour)
                    date_str = full_date.strftime("%d/%m/%Y")
                except:
                    continue
                if str(ws.cell(row=row, column=matin_col).value).strip().upper() == "X":
                    dispo[nom][date_str].add("AM")
                if str(ws.cell(row=row, column=am_col).value).strip().upper() == "X":
                    dispo[nom][date_str].add("PM")
    wb.close()
    return dispo

def charger_heures(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    heures = {}
    for nom in wb.sheetnames:
        ws = wb[nom]
        heures[nom] = {}
        for row in range(4, ws.max_row + 1):
            classe = ws.cell(row=row, column=1).value
            h = ws.cell(row=row, column=2).value
            if classe and isinstance(h, (int, float)):
                heures[nom][classe] = h
    wb.close()
    return heures

def attribuer_cours(cours, dispos, heures_init):
    heures_restantes = {prof: heures_init[prof].copy() for prof in heures_init}
    affectations = defaultdict(list)

    for date, moment, classe, feuille, row, col in cours:
        candidats = []
        for prof in dispos:
            if moment in dispos[prof].get(date, set()) and classe in heures_restantes[prof]:
                if heures_restantes[prof][classe] >= 4:
                    candidats.append((prof, heures_restantes[prof][classe]))
        if not candidats:
            continue
        candidats.sort(key=lambda x: x[1])
        choisi = candidats[0][0]
        affectations[choisi].append((date, moment, classe, feuille, row, col))
        heures_restantes[choisi][classe] -= 4
        dispos[choisi][date].discard(moment)
    return affectations

def enregistrer_affectations_excel(fichier_source, affectations, fichier_sortie):
    wb = openpyxl.load_workbook(fichier_source)
    for prof, cours in affectations.items():
        for date, moment, classe, feuille, row, col in cours:
            ws = wb[feuille]
            ws.cell(row=row, column=col).value = prof
    wb.save(fichier_sortie)

def generer_excel(fichier_heures, fichier_prof, fichier_mois, fichier_sortie):
    try:
        heures = charger_heures(fichier_heures)
        dispos = charger_dispos(fichier_prof)
        cours = charger_cours(fichier_mois)
        affectations = attribuer_cours(cours, dispos, heures)
        enregistrer_affectations_excel(fichier_mois, affectations, fichier_sortie)
    except Exception as e:
        st.error(f"❌ Erreur dans la génération : {e}")

# === INTERFACE STREAMLIT ===
st.set_page_config(page_title="📅 Générateur LODIMA", layout="centered")
st.title("📅 Générateur d'emploi du temps - LODIMA")

uploaded_mois = st.file_uploader("📂 Importer le fichier Mois.xlsx", type="xlsx")
uploaded_prof = st.file_uploader("👤 Importer le fichier Prof.xlsx", type="xlsx")
uploaded_heures = st.file_uploader("⏱️ Importer le fichier Heure.xlsx", type="xlsx")

if uploaded_mois and uploaded_prof and uploaded_heures:
    if st.button("🚀 Générer le fichier"):
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_mois, \
             NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_prof, \
             NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_heures, \
             NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_result:

            tmp_mois.write(uploaded_mois.read())
            tmp_prof.write(uploaded_prof.read())
            tmp_heures.write(uploaded_heures.read())

            tmp_mois.flush()
            tmp_prof.flush()
            tmp_heures.flush()

            generer_excel(
                fichier_heures=tmp_heures.name,
                fichier_prof=tmp_prof.name,
                fichier_mois=tmp_mois.name,
                fichier_sortie=tmp_result.name
            )

            with open(tmp_result.name, "rb") as f:
                st.success("✅ Fichier généré avec succès !")
                st.download_button("📥 Télécharger le fichier Excel", f, file_name="Mois_avec_profs.xlsx")

else:
    st.info("📁 Veuillez importer les trois fichiers requis.")
